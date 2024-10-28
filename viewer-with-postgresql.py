import os
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from openpyxl import Workbook
import pandas as pd
import psycopg2
from psycopg2.extras import Json
import bcrypt

class JsonViewerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JSON Viewer - PostgreSQL")
        self.root.geometry("800x600")

        # Database connection
        self.conn = psycopg2.connect(
            dbname="dbname",
            user="dbuser",
            password="****",
            host="****",
            port="****"
        )
        self.cursor = self.conn.cursor()
        self.create_tables()
        self.add_default_admin()

        #frame layout
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=1)

        #toolbar
        self.create_toolbar()

        self.current_user = None

        #paned window
        self.paned_window = tk.PanedWindow(self.main_frame, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill=tk.BOTH, expand=1)

        #left frame
        self.list_frame = tk.Frame(self.paned_window, width=200)
        self.list_frame.pack_propagate(False)

        #search box
        self.search_var = tk.StringVar()
        self.search_box = tk.Entry(self.list_frame, textvariable=self.search_var)
        self.search_box.pack(fill=tk.X, padx=5, pady=5)
        self.search_box.bind("<KeyRelease>", self.update_file_list)

        #listbox frame and scrollbars
        self.list_scroll_y = tk.Scrollbar(self.list_frame, orient=tk.VERTICAL)
        self.list_scroll_x = tk.Scrollbar(self.list_frame, orient=tk.HORIZONTAL)
        self.file_listbox = tk.Listbox(self.list_frame, yscrollcommand=self.list_scroll_y.set,
                                       xscrollcommand=self.list_scroll_x.set)
        self.file_listbox.bind("<<ListboxSelect>>", self.display_json_content)
        self.list_scroll_y.config(command=self.file_listbox.yview)
        self.list_scroll_x.config(command=self.file_listbox.xview)
        self.list_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.list_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.file_listbox.pack(fill=tk.BOTH, expand=1)

        self.paned_window.add(self.list_frame)

        #treeview frame
        self.tree_frame = tk.Frame(self.paned_window)
        self.tree_scroll_y = tk.Scrollbar(self.tree_frame, orient=tk.VERTICAL)
        self.tree_scroll_x = tk.Scrollbar(self.tree_frame, orient=tk.HORIZONTAL)
        self.tree = ttk.Treeview(self.tree_frame, yscrollcommand=self.tree_scroll_y.set,
                                 xscrollcommand=self.tree_scroll_x.set)
        self.tree_scroll_y.config(command=self.tree.yview)
        self.tree_scroll_x.config(command=self.tree.xview)
        self.tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=1)

        self.paned_window.add(self.tree_frame)

        #buttons
        self.import_button = tk.Button(self.root, text="Import JSON Files", command=self.import_json_files)
        self.import_button.pack(side=tk.BOTTOM, pady=5)

        self.export_button = tk.Button(self.root, text="Export to Excel", command=self.export_to_excel)
        self.export_button.pack(side=tk.BOTTOM, pady=5)

        self.minimal_report_button = tk.Button(self.root, text="Minimal Report", command=self.create_minimal_report)
        self.minimal_report_button.pack(side=tk.BOTTOM, pady=5)

        # User authentication
        self.current_user = None
        self.show_login_dialog()

    def create_tables(self):
        try:
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS json_files (
                    id SERIAL PRIMARY KEY,
                    filename TEXT UNIQUE NOT NULL,
                    content JSONB NOT NULL
                )
            """)
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    is_default BOOLEAN DEFAULT FALSE
                )
            """)
            self.conn.commit()
            print("Tables created successfully.")
        except Exception as e:
            print(f"Error creating tables: {e}")
            self.conn.rollback()

    def add_default_admin(self):
        hashed_password = bcrypt.hashpw(b"123", bcrypt.gensalt())
        try:
            self.cursor.execute("""
                INSERT INTO users (username, password, is_default)
                VALUES (%s, %s, %s)
                ON CONFLICT (username) DO NOTHING
            """, ("admin", hashed_password.decode('utf-8'), True))
            self.conn.commit()
            print("Default admin added successfully.")
        except psycopg2.Error as e:
            print(f"Error adding default admin: {e}")
            self.conn.rollback()

    def create_toolbar(self):
        toolbar = tk.Frame(self.root, bd=1, relief=tk.RAISED)
        toolbar.pack(side=tk.TOP, fill=tk.X)

        login_button = tk.Button(toolbar, text="Login", command=self.show_login_dialog)
        login_button.pack(side=tk.LEFT, padx=2, pady=2)

        create_user_button = tk.Button(toolbar, text="Create User", command=self.show_create_user_dialog)
        create_user_button.pack(side=tk.LEFT, padx=2, pady=2)

        logout_button = tk.Button(toolbar, text="Logout", command=self.logout)
        logout_button.pack(side=tk.LEFT, padx=2, pady=2)

    def show_login_dialog(self):
        username = simpledialog.askstring("Login", "Enter username:")
        if username:
            password = simpledialog.askstring("Login", "Enter password:", show='*')
            if password:
                self.login(username, password)

    def show_create_user_dialog(self):
        username = simpledialog.askstring("Create User", "Enter new username:")
        if username:
            password = simpledialog.askstring("Create User", "Enter password:", show='*')
            if password:
                self.create_user(username, password)

    def login(self, username, password):
        self.cursor.execute("SELECT password, is_default FROM users WHERE username = %s", (username,))
        result = self.cursor.fetchone()
        if result and bcrypt.checkpw(password.encode('utf-8'), result[0].encode('utf-8')):
            self.current_user = username
            messagebox.showinfo("Login", f"Welcome, {username}!")

            #if it's the default admin user, disable it after successful login
            if result[1]: 
                self.disable_default_admin()

            self.load_json_files()
        else:
            messagebox.showerror("Login", "Invalid username or password")

    def logout(self):
        self.current_user = None
        self.file_listbox.delete(0, tk.END)
        self.tree.delete(*self.tree.get_children())
        messagebox.showinfo("Logout", "You have been logged out")

    def create_user(self, username, password):
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        try:
            self.cursor.execute("INSERT INTO users (username, password) VALUES (%s, %s)",
                                (username, hashed_password.decode('utf-8')))
            self.conn.commit()
            messagebox.showinfo("Create User", f"User {username} created successfully")
        except psycopg2.IntegrityError:
            self.conn.rollback()
            messagebox.showerror("Create User", f"Username {username} already exists")

    def disable_default_admin(self):
        try:
            self.cursor.execute("DELETE FROM users WHERE username = 'admin' AND is_default = TRUE")
            self.conn.commit()
            print("Default admin user has been disabled.")
        except psycopg2.Error as e:
            print(f"Error disabling default admin: {e}")

    def load_json_files(self):
        """load file names from the database"""
        if not self.current_user:
            messagebox.showerror("Error", "Please log in to view files")
            return
        self.cursor.execute("SELECT filename FROM json_files")
        self.json_files = [row[0] for row in self.cursor.fetchall()]
        self.update_file_list()

    def update_file_list(self, event=None):
        """update file list based on search"""
        search_term = self.search_var.get().lower()
        self.file_listbox.delete(0, tk.END)
        for file in self.json_files:
            if search_term in file.lower():
                self.file_listbox.insert(tk.END, file)

    def import_json_files(self):
        """import files into the database"""
        if not self.current_user:
            messagebox.showerror("Error", "Please log in to import files")
            return
        file_paths = filedialog.askopenfilenames(filetypes=[("JSON files", "*.json")])
        for file_path in file_paths:
            filename = os.path.basename(file_path)
            with open(file_path, 'r') as file:
                json_data = json.load(file)

            self.cursor.execute(
                "INSERT INTO json_files (filename, content) VALUES (%s, %s) ON CONFLICT (filename) DO UPDATE SET content = EXCLUDED.content",
                (filename, Json(json_data))
            )
        self.conn.commit()
        self.load_json_files()

    def display_json_content(self, event):
        """display file content in the tree view"""
        if not self.current_user:
            messagebox.showerror("Error", "Please log in to view file contents")
            return
        selected_file = self.file_listbox.get(tk.ACTIVE)
        self.cursor.execute("SELECT content FROM json_files WHERE filename = %s", (selected_file,))
        json_data = self.cursor.fetchone()[0]

        self.tree.delete(*self.tree.get_children())
        self.populate_tree('', json_data)

    def populate_tree(self, parent, json_data):
        """populate treeview"""
        if isinstance(json_data, dict):
            for key, value in json_data.items():
                node_id = self.tree.insert(parent, 'end', text=str(key), open=True)
                self.populate_tree(node_id, value)
        elif isinstance(json_data, list):
            for idx, value in enumerate(json_data):
                node_id = self.tree.insert(parent, 'end', text=f"Item {idx}", open=True)
                self.populate_tree(node_id, value)
        else:
            self.tree.insert(parent, 'end', text=str(json_data))

    def export_to_excel(self):
        """export all Excel"""
        if not self.current_user:
            messagebox.showerror("Error", "Please log in to export files")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "JSON Data"

        row_num = 1
        self.cursor.execute("SELECT filename, content FROM json_files")
        for filename, json_data in self.cursor.fetchall():
            ws.cell(row=row_num, column=1, value=f"File: {filename}")
            row_num = self.write_json_to_sheet(ws, json_data, row_num + 1)
            row_num += 1 

        # Save the Excel file
        excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx")])
        if excel_path:
            wb.save(excel_path)
            messagebox.showinfo("Success", "Excel file created successfully!")

    def write_json_to_sheet(self, sheet, json_data, start_row, start_col=1):
        """Write content to Excel"""
        if isinstance(json_data, dict):
            for key, value in json_data.items():
                sheet.cell(row=start_row, column=start_col, value=key)
                start_row = self.write_json_to_sheet(sheet, value, start_row + 1, start_col + 1)
        elif isinstance(json_data, list):
            for idx, value in enumerate(json_data):
                sheet.cell(row=start_row, column=start_col, value=f"Item {idx}")
                start_row = self.write_json_to_sheet(sheet, value, start_row + 1, start_col + 1)
        else:
            sheet.cell(row=start_row, column=start_col, value=json_data)
        return start_row

    def create_minimal_report(self):
        """create a minimal report"""
        if not self.current_user:
            messagebox.showerror("Error", "Please log in to create a report")
            return
        all_data = []

        self.cursor.execute("SELECT filename, content FROM json_files")
        for filename, data in self.cursor.fetchall():
            filename_parts = filename.split('_')
            row = {
                'ComputerName': filename_parts[0] if len(filename_parts) > 0 else '',
                'DomainName': filename_parts[1] if len(filename_parts) > 1 else '',
                'DateTime': filename_parts[2].split('.')[0] if len(filename_parts) > 2 else ''
            }

            #CPU
            cpu_list = data.get('CPU', [{}])
            for cpu in cpu_list:
                row.update({
                    'CPU_Name': cpu.get('Name', ''),
                    'CPU_SocketDesignation': cpu.get('SocketDesignation', '')
                })

            #Motherboard
            motherboard_list = data.get('Motherboard', [{}])
            for motherboard in motherboard_list:
                row.update({
                    'MB_Manufacturer': motherboard.get('Manufacturer', ''),
                    'MB_Product': motherboard.get('Product', '')
                })

            #Memory Modules
            memory_modules = data.get('MemoryModules', [])
            for i, module in enumerate(memory_modules):
                prefix = f'MemoryModule_{i + 1}_'
                row.update({
                    f'{prefix}Manufacturer': module.get('Manufacturer', ''),
                    f'{prefix}Capacity': module.get('Capacity', '')
                })

            #Printers
            printers = data.get('Printers', [])
            for i, printer in enumerate(printers):
                if printer.get('PrinterStatus') == 'Online':
                    prefix = f'Printer_{i + 1}_'
                    row.update({
                        f'{prefix}DriverName': printer.get('DriverName')
                    })


            #WIA Devices
            wia_devices = data.get('WIADevices', [])
            #Combine WIA device names into a single column
            wia_device_names = [device.get('Name', device) if isinstance(device, dict) else device for device in
                                wia_devices]
            row['WIADevices'] = ', '.join(wia_device_names)

            #DVD/CD-ROM
            dvd_drives = data.get('DVD/CD-ROM', [])
            for i, drive in enumerate(dvd_drives):
                prefix = f'DVDDrive_{i + 1}_'
                row.update({
                    f'{prefix}Caption': drive.get('Caption'),
                    f'{prefix}Id': drive.get('Id')
                })

            #Disks
            disks = data.get('Disks', [])
            for i, disk in enumerate(disks):
                prefix = f'Disk_{i + 1}_'
                row.update({
                    f'{prefix}Model': disk.get('Model', ''),
                    f'{prefix}Size': disk.get('Size', '')
                })

            #Operating System
            os_data = data.get('OperatingSystem', [{}])
            for os_info in os_data:
                row.update({
                    'OS_Caption': os_info.get('Caption', ''),
                    'OS_InstallDate': os_info.get('InstallDate', '')
                })

            #Network Adapters
            network_adapters = data.get('NetworkAdapters', [])
            for i, adapter in enumerate(network_adapters):
                prefix = f'NetworkAdapter_{i + 1}_'
                row.update({
                    f'{prefix}Description': adapter.get('Description', ''),
                    f'{prefix}MACAddress': adapter.get('MACAddress', ''),
                    f'{prefix}IPv4': adapter.get('IPv4', '')
                })

            #Windows SID
            row['WindowsSID'] = data.get('WindowsSID')

            all_data.append(row)

        df = pd.DataFrame(all_data)
        df = df.reindex(sorted(df.columns), axis=1)

        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                   filetypes=[("Excel files", "*.xlsx")])
        if output_path:
            df.to_excel(output_path, index=False)
            messagebox.showinfo("Success", "Minimal report created successfully!")


    def __del__(self):
        """Close the database connection when the object is destroyed."""
        if hasattr(self, 'conn'):
            self.conn.close()

if __name__ == "__main__":
    root = tk.Tk()
    app = JsonViewerApp(root)
    root.mainloop()
