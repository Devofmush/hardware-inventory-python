import os
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook

class JsonViewerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JSON Viewer")
        self.root.geometry("800x600")
        # frame layout
        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=1)
        # paned window
        self.paned_window = tk.PanedWindow(self.main_frame, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill=tk.BOTH, expand=1)
        # Left frame
        self.list_frame = tk.Frame(self.paned_window, width=200)
        self.list_frame.pack_propagate(False)  # Prevents shrinking to fit contents
        # search box
        self.search_var = tk.StringVar()
        self.search_box = tk.Entry(self.list_frame, textvariable=self.search_var)
        self.search_box.pack(fill=tk.X, padx=5, pady=5)
        self.search_box.bind("<KeyRelease>", self.update_file_list)
        # Listbox frame
        self.list_scroll_y = tk.Scrollbar(self.list_frame, orient=tk.VERTICAL)
        self.list_scroll_x = tk.Scrollbar(self.list_frame, orient=tk.HORIZONTAL)
        self.file_listbox = tk.Listbox(self.list_frame, yscrollcommand=self.list_scroll_y.set, xscrollcommand=self.list_scroll_x.set)
        self.file_listbox.bind("<<ListboxSelect>>", self.display_json_content)
        self.list_scroll_y.config(command=self.file_listbox.yview)
        self.list_scroll_x.config(command=self.file_listbox.xview)
        self.list_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.list_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.file_listbox.pack(fill=tk.BOTH, expand=1)
        self.paned_window.add(self.list_frame)
        self.tree_frame = tk.Frame(self.paned_window)
        self.tree_scroll_y = tk.Scrollbar(self.tree_frame, orient=tk.VERTICAL)
        self.tree_scroll_x = tk.Scrollbar(self.tree_frame, orient=tk.HORIZONTAL)
        self.tree = ttk.Treeview(self.tree_frame, yscrollcommand=self.tree_scroll_y.set, xscrollcommand=self.tree_scroll_x.set)
        self.tree_scroll_y.config(command=self.tree.yview)
        self.tree_scroll_x.config(command=self.tree.xview)
        self.tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=1)
        self.paned_window.add(self.tree_frame)
        self.export_button = tk.Button(self.root, text="Export to Excel", command=self.export_to_excel)
        self.export_button.pack(side=tk.BOTTOM, pady=10)
        self.json_files = []
        self.load_json_files()

    def load_json_files(self):
        """Load file names from the 'json_files' dir"""
        self.json_dir = os.path.join(os.getcwd(), "json_files")
        if not os.path.exists(self.json_dir):
            os.makedirs(self.json_dir)
        self.json_files = [f for f in os.listdir(self.json_dir) if f.endswith(".json")]
        self.update_file_list()

    def update_file_list(self, event=None):
        """Update list on search"""
        search_term = self.search_var.get().lower()
        self.file_listbox.delete(0, tk.END)
        for file in self.json_files:
            if search_term in file.lower():
                self.file_listbox.insert(tk.END, file)

    def display_json_content(self, event):
        """Display content in the tree view"""
        selected_file = self.file_listbox.get(tk.ACTIVE)
        file_path = os.path.join(self.json_dir, selected_file)
        with open(file_path, 'r') as file:
            json_data = json.load(file)

        self.tree.delete(*self.tree.get_children())
        self.populate_tree('', json_data)

    def populate_tree(self, parent, json_data):
        """populate treeview"""
        if isinstance(json_data, dict):
            for key, value in json_data.items():
                node_id = self.tree.insert(parent, 'end', text=str(key), open=True)
                self.populate_tree(node_id, value)
        elif isinstance(json_data, list):
            for idx, value in enumerate(json_data):
                node_id = self.tree.insert(parent, 'end', text=f"Item {idx}", open=True)
                self.populate_tree(node_id, value)
        else:
            self.tree.insert(parent, 'end', text=str(json_data))

    def export_to_excel(self):
        """Export all"""
        wb = Workbook()
        ws = wb.active
        ws.title = "JSON Data"

        row_num = 1
        for json_file in os.listdir(self.json_dir):
            if json_file.endswith(".json"):
                file_path = os.path.join(self.json_dir, json_file)
                with open(file_path, 'r') as file:
                    json_data = json.load(file)

                # file name in the first column
                ws.cell(row=row_num, column=1, value=f"File: {json_file}")
                row_num = self.write_json_to_sheet(ws, json_data, row_num + 1)
                row_num += 1 

        # excel file
        excel_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx")])
        if excel_path:
            wb.save(excel_path)
            messagebox.showinfo("Success", "Excel file created successfully!")

    def write_json_to_sheet(self, sheet, json_data, start_row, start_col=1):
        """Write to an Excel sheet."""
        if isinstance(json_data, dict):
            for key, value in json_data.items():
                sheet.cell(row=start_row, column=start_col, value=key)
                start_row = self.write_json_to_sheet(sheet, value, start_row + 1, start_col + 1)
        elif isinstance(json_data, list):
            for idx, value in enumerate(json_data):
                sheet.cell(row=start_row, column=start_col, value=f"Item {idx}")
                start_row = self.write_json_to_sheet(sheet, value, start_row + 1, start_col + 1)
        else:
            sheet.cell(row=start_row, column=start_col, value=json_data)
        return start_row

if __name__ == "__main__":
    root = tk.Tk()
    app = JsonViewerApp(root)
    root.mainloop()
